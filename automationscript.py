from winsound import Beep
from time import sleep
import openpyxl

wynik = 0
dyski = ["BRAK", "SEAGATE BARRACUDA", 'WD BLUE', "WD CAVLAR", "MAXTOR","TOSHIBA","HITACHI"]
wielkosc = ["BRAK", "500 GB", "250 GB", "40 GB", "1 TB", "120 GB"]
starts = ["CZ", "08", "PC", "PK"]


def gettingsmart():
    global smartdict
    smartdict = {}
    for x in open("pattern.txt", "r").readlines():
        x = x.split(",")
        smartdict[x[0]] = [x[1], x[2].rstrip()]
    stagezero()


def stagezero():
    global curbox, excel, PCs, CZCs, PKOs, exc_path
    curbox = input("\nPodaj nazwę boxu (np 'BOX 13'): \n").upper()
    completed = False
    while not completed:
        exc_path = input("Podaj nazwe/sciezke dzisiejszego pliku excela (.xlsx, foreslashs /): ")
        try:
            excel = openpyxl.load_workbook(exc_path)
            completed = True
        except openpyxl.utils.exceptions.InvalidFileException:
            Beep(500, 1000)
            print("Plik o zlym rozszerzeniu\n")
        except FileNotFoundError:
            Beep(500, 1000)
            print("Nie znaleziono pliku w danej sciezce\n")
    PCs = []
    for row in excel["Arkusz1"].iter_rows(min_row=1, min_col=2, max_col=2, values_only=True):
        PCs.append(row[0])
    CZCs = []
    for row in excel["Arkusz1"].iter_rows(min_row=1, min_col=4, max_col=4, values_only=True):
        CZCs.append(row[0])
    PKOs = []
    for row in excel["Arkusz1"].iter_rows(min_row=1, min_col=3, max_col=3, values_only=True):
        PKOs.append(row[0])
    stageone()


# znalezienie assetow i indeksu komputera
def stageone():
    global curbox, wynik, excel
    inp = input("\nEKRAN STARTOWY\nObecny wynik: "+ str(wynik) +
                "\nPodaj argument lub rozpocznij skanowanie (PC/SN/PKO)\nObecny box: "+
                curbox + "\n\nsave - zapisuje plik\nreload - przeladowuje plik\nBOX + nr - ustawia domyslny pojemnik \nend - konczy prace \n")
    if inp.lower() == "end":
        print("Twoj dzisiejszy wynik: " + str(wynik))
        print("Poczekaj...\n")
        try:
            excel.save(exc_path)
        except PermissionError:
            Beep(500, 1000)
            print("Niepowodzenie\nZamknij excela z plikiem\n")
            stageone()
        sleep(5)
        raise SystemExit
    elif inp.lower() == "save":
        print("Poczekaj...\n")
        try:
            excel.save(exc_path)
        except PermissionError:
            Beep(500, 1000)
            print("Niepowodzenie\nZamknij excela z plikiem\n")
        stageone()
    elif inp[:3].upper() == "BOX":
        curbox = inp
        stageone()
    elif inp.lower() == "reload":
        excel = openpyxl.load_workbook(exc_path)
        print("gotowe\n")
        stageone()
    elif len(inp) == 8 and inp[:2] == "PC":
        try:
            ind = PCs.index(inp)
            stagetwo(ind)
        except ValueError:
            print("Nie znaleziono sprzetu dla tego PC\n")
            Beep(500, 1000)
            stageone()
    elif len(inp) == 10 and inp[:3] == "CZC":
        try:
            ind = CZCs.index(inp)
            stagetwo(ind)
        except ValueError:
            print("Nie znaleziono sprzetu dla tego CZC\n")
            Beep(500, 1000)
            stageone()
    elif len(inp) == 11 and inp[:4] == "PKO-" or len(inp) == 7 and inp[:2] == "08":
        if len(inp) == 7:
            inp = "PKO-" + inp
        try:
            ind = PKOs.index(inp)
            stagetwo(ind)
        except ValueError:
            print("Nie znaleziono sprzetu dla tego PKO\n")
            Beep(500, 1000)
            stageone()
    else:
        print("Nie rozumiem. Napisales: '" + inp + "'\n")
        Beep(500, 1000)
        stageone()


# wczytanie SN dysku
def stagetwo(ind):
    global excel
    print("\n\nObecny komputer to:")
    assets = []
    for col in excel["Arkusz1"].iter_cols(min_row=ind + 1, min_col=2, max_col=4, max_row=ind + 1, values_only=True):
        print(col[0])
        assets.append(col[0])
    inp = input(
        "Mozesz zeskanowac ponownie ktorys z powyzszych assetow, aby porownac\nAby wrocic, wyslij Z\nSczytaj SN dysku: ")
    if inp[:2] in starts:
        if len(inp) == 7:
            inp = "PKO-" + inp
        if inp in assets:
            print(inp + " zgadza sie z wczytanymi assetami\n")
        else:
            print(inp + " NIE zgadza sie z wczytanymi assetami\n")
            Beep(500, 1000)
        stagetwo(ind)
    elif inp.upper() == "Z":
        print("Powrot do ekranu poczatkowego\n")
        stageone()
    elif inp.upper() == "R":
        print("Powrot do ekranu poczatkowego\n")
        stageone()
    else:
        smartsn(inp, ind)


# okreslenie nazwy dysku
def stagethree(sn, ind):
    global dyski
    print("\n\nZeskanowany SN dysku: " + sn + "\nAby ponownie zeskanowac, wyslij Z")
    for x in range(len(dyski)):
        print(str(x) + " - " + dyski[x])
    name = input("Podaj nazwe: ")
    if name.upper() == "Z":
        print("Powrot do wczytania SN dysku\n")
        stagetwo(ind)
    elif name.upper() == "R":
        print("Powrot do ekranu poczatkowego\n")
        stageone()
    try:
        name = dyski[int(name)]
    except ValueError:
        pass
    except IndexError:
        print("Nie rozumiem. Napisales: '" + name + "'\n")
        Beep(500, 1000)
        stagethree(sn, ind)
    stagefour(name, sn, ind)


# okreslenie wielkosci dysku
def stagefour(name, sn, ind):
    global wielkosc
    print("\n\nNazwa dysku: " + name + "\nAby ponownie wprowadzic nazwe, wyslij Z")
    for x in range(len(wielkosc)):
        print(str(x) + " - " + wielkosc[x])
    size = input("Podaj wielkosc: ")
    if size.upper() == "Z":
        print("Powrot do podania nazwy\n")
        stagethree(sn, ind)
    elif size.upper() == "R":
        print("Powrot do ekranu poczatkowego\n")
        stageone()
    try:
        size = wielkosc[int(size)]
    except ValueError:
        pass
    except IndexError:
        print("Nie rozumiem. Napisales: '" + size + "'\n")
        Beep(500, 1000)
        stagefour(name, sn, ind)
    stagefive(size, name, sn, ind)


# okreslenie ramu
def stagefive(size, name, sn, ind):
    print("\n\nWielkosc dysku: " + size + "\nAby ponownie wprowadzic wielkosc, wyslij Z")
    ram = input("Jesli brak ramu wyslij 0, jesli obecny, kliknij enter: ")
    if ram.upper() == "Z":
        print("Powrot do podania wielkosci\n")
        stagefour(name, sn, ind)
    elif ram.upper() == "R":
        print("Powrot do ekranu poczatkowego\n")
        stageone()
    elif ram == "0":
        ram = False
        stagesix(ram, size, name, sn, ind)
    elif ram == "":
        ram = True
        stagesix(ram, size, name, sn, ind)
    else:
        print("Nie rozumiem. Napisales: '" + ram + "'\n")
        Beep(500, 1000)
        stagefive(size, name, sn, ind)


# okreslenie boxu
def stagesix(ram, size, name, sn, ind):
    global curbox
    if not ram:
        isram = "Brak ramu"
    else:
        isram = "Ram obecny"
    print("\n\n"+isram + "\nAby ponownie wprowadzic ram, wyslij Z\nDomyslny: " + curbox)
    box = input("Jesli chcesz inny box, podaj jego nazwe, inaczej wcisnij enter: ")
    if box.upper() == "Z":
        print("Powrot do podania ramu\n")
        stagefive(size, name, sn, ind)
    elif box.upper() == "R":
        print("Powrot do ekranu poczatkowego\n")
        stageone()
    elif box == "":
        writing(curbox, ram, size, name, sn, ind)
    else:
        writing(box, ram, size, name, sn, ind)


# wpisanie do excela
def writing(box, ram, size, name, sn, ind):
    global wynik, PKOs, excel, CZCs, exc_path
    if not ram:
        isram = "Brak ramu"
    else:
        isram = "Ram obecny"
    curpko = PKOs[ind]
    print(curpko, CZCs[ind], sn, name, size, isram, box)
    wr = input("\n\nAby zmienic box wyslij Z, aby wpisac do excela wcisnij enter: ")
    if wr.upper() == "Z":
        print("Powrot do podania boxa\n")
        stagesix(ram, size, name, sn, ind)
    elif wr.upper() == "R":
        print("Powrot do ekranu poczatkowego\n")
        stageone()
    elif wr == "":
        log = open("logs","a")
        log.write(str([name, size, sn, box, curpko, isram, CZCs[ind]])+"\n")
        log.close()
        for x, y in zip([name, size, sn, box, curpko], [10, 11, 12, 13, 14]):
            excel["Arkusz1"].cell(row=ind + 1, column=y, value=x)
        if ram:
            sht = excel["PC_PKO_Z_RAMEM"]
        else:
            sht = excel["PC_PKO_BEZ_RAM"]
        for z in range(1, 1000):
            if sht.cell(row=z, column=1).value is None:
                sht.cell(row=z, column=1, value=CZCs[ind])
                break
        if wynik%20==1:
            completed = False
            while not completed:
                try:
                    print("Poczekaj...")
                    excel.save(exc_path[:-5]+"_"+str(wynik)+".xlsx")
                    completed = True
                except PermissionError:
                    Beep(500, 1000)
                    print("Niepowodzenie\nZamknij excela z plikiem\n")
        completed = False
        while not completed:
            try:
                print("Poczekaj...")
                excel.save(exc_path)
                completed = True
            except PermissionError:
                Beep(500, 1000)
                print("Niepowodzenie\nZamknij excela z plikiem\n")
        wynik += 1
        stageone()
    else:
        Beep(500, 1000)
        print("Nie rozumiem. Napisales: '" + wr + "'\n")
        writing(box,ram, size, name, sn, ind)


def smartsn(sn, ind):
    if sn[:3] in smartdict.keys():
        guess = smartdict[sn[:3]]
        print(guess)
        match = input("\n\nJesli to dobry dysk, kliknij enter, jesli zle, wyslij zero. Aby ponownie szczytac SN wyslij Z: ")
        if match.upper() == "Z":
            print("Powrot do wczytania SN\n")
            stagetwo(ind)
        elif match.upper() == "R":
            print("Powrot do ekranu poczatkowego\n")
            stageone()
        elif match == "0":
            stagethree(sn, ind)
        elif match == "":
            stagefive(guess[1], guess[0], sn, ind)
        else:
            print("Nie rozumiem. Napisales: '" + match + "'\n")
            Beep(500, 1000)
            smartsn(sn, ind)
    stagethree(sn, ind)


gettingsmart()
